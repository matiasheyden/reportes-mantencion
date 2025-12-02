"""Exporta todas las hojas de BBDD_MANTENCION.xlsm a CSV.

Uso: ejecutar este módulo desde el venv del workspace.
Genera los CSV en la carpeta `csv_exports/` dentro del workspace.
"""
from pathlib import Path
import os
import re
import sys


def _safe_filename(name: str) -> str:
    # Reemplaza caracteres no permitidos por guiones bajos
    name = name.strip()
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    name = re.sub(r"\s+", "_", name)
    return name


def export_all_sheets(xls_path: Path, out_dir: Path) -> dict:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = {"written": [], "errors": []}

    # Intentar usar pandas primero
    try:
        import pandas as pd
        # openpyxl engine for xlsm
        sheets = pd.read_excel(xls_path, sheet_name=None, engine="openpyxl")
        for sheet_name, df in sheets.items():
            fname = out_dir / (f"{_safe_filename(sheet_name)}.csv")
            try:
                df.to_csv(fname, index=False, encoding="utf-8-sig")
                summary["written"].append(str(fname))
            except Exception as e:
                summary["errors"].append({"sheet": sheet_name, "error": str(e)})
        return summary
    except Exception:
        # fallback a openpyxl si pandas no está presente
        pass

    try:
        from openpyxl import load_workbook
        import csv

        wb = load_workbook(filename=str(xls_path), data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            fname = out_dir / (f"{_safe_filename(sheet_name)}.csv")
            try:
                with open(fname, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        # Convertir objetos a str para evitar errores de serialización
                        writer.writerow(["" if v is None else str(v) for v in row])
                summary["written"].append(str(fname))
            except Exception as e:
                summary["errors"].append({"sheet": sheet_name, "error": str(e)})
        return summary
    except Exception as e:
        summary["errors"].append({"error": str(e)})
        return summary


def main():
    workspace = Path(__file__).parent
    xls = workspace / "BBDD_MANTENCION.xlsm"
    if not xls.exists():
        print(f"ERROR: no se encontró el archivo: {xls}")
        sys.exit(1)

    out_dir = workspace / "csv_exports"
    print(f"Leyendo: {xls}")
    print(f"Exportando hojas a: {out_dir}")

    summary = export_all_sheets(xls, out_dir)

    print("--- Resumen ---")
    print(f"Archivos escritos: {len(summary.get('written',[]))}")
    for w in summary.get("written", []):
        print(f" - {w}")
    if summary.get("errors"):
        print(f"Errores ({len(summary['errors'])}):")
        for err in summary["errors"]:
            print(" -", err)


if __name__ == "__main__":
    main()
